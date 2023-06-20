import xlrd
import xlwt
import openpyxl
import json
from xlutils.copy import copy
# 指定工作簿路径
excel_name=('F:/Excel/测试工作表.xlsx')

# 对表一进行操作
"""
# 获取指定对象
work_book = xlrd.open_workbook(excel_name)
# 获取工作簿所有sheet表对象名称
sheets = work_book.sheet_names()
# 获取sheet1表
sheet1 = work_book.sheet_by_index(0)
# 在sheet1表中，获取单个单元格进行操作

cell_0_0=sheet1.cell(0,0)
print(type(cell_0_0))
print("第一个单元格信息",cell_0_0)

# 在sheet1表中，获取行进行操作

# 获取sheet表对象有效行数
row_sum = sheet1.nrows
print("有效行数:",row_sum)

# 获取指定行的长度
row_len = sheet1.row_len(0)
print("指定行的长度:",row_len)

# 获取所有行
rows = sheet1.get_rows()
for row in rows:
    print(row)


# 获取列操作
# 统计有效列数
cols_sum=sheet1.ncols
print("有效列数为：",cols_sum)
# 获取某一列的值
cols_value=sheet1.col_values(0)
print("第一列的值为：",cols_value)
# 获取某一列的数据类型
cols_type=sheet1.col_types(0)
print("第一列的数据类型：",cols_type)
# 获取有效列
cols = sheet1.col_values(0)

for col in cols:
    print(col)
"""

"""
对表进行覆写操作
"""
"""
# def WriteExcel(path,sheet_name,write_dict):
#     '''
#     #   :param sheet_name: 需要改写的sheet_name
#     #   :param path: 工作表的路径
#     #   :return: 
#     '''
#     work_book = openpyxl.Workbook()
#     sheet = work_book.active
#     sheet.title = sheet_name
#     key_list = []
#     value_list = []
#     for k,v in write_dict.items():
#         key_list.append(k)
#         value_list.append(v)
#         # 定义第一行表头
#         for i in range(0,len(key_list)):
#             sheet.cell(row=1,column=i+1,value=key_list[i])
#             # 定义以下的的参数
#             for j in range(0,len(value_list)):
#                 sheet.cell(row=i+2,column=j+1,value=value_list[j])
#     work_book.save(path)
#     print("xlsx格式表格【覆写】写入数据成功")
# if __name__=='__main__':
#     # 定义要写入的excel表的绝对路径
#     path='F:/Excel/测试工作表.xlsx'
#     # 定义要写入的行和列的值
#     write_dict={
#         "请求方法":"post",
#         "请求参数":"测试参数",
#         "返回参数":"测试参数",
#         "预期结果":"successful",
#         "实际结果":"successful",
#     }
#     # 定义excel的sheet_name
#     sheet_name = 'Sheet1'
#     WriteExcel(path,sheet_name,write_dict)
"""

"""
对表二进行覆写操作,进使用xlrd和xlwt实现
"""

save_path=('F:/Excel/student.xlsx')
# 获取读取对象
work_book = xlrd.open_workbook(save_path)
# 获取写入对象
tc_wb = xlwt.Workbook()

# 指定要添加的表
sheet2_write = tc_wb.add_sheet('Sheet1')

# 获取表二表单
sheet2 = work_book.sheet_by_name('Sheet2')

# 获取Sheet2有效行,列
rows = sheet2.nrows
cols = sheet2.ncols
# 获取Sheet2,每一个单元格的数据
for row in range(0,rows):
    for col in range(0,cols):
        cell_vlue=sheet2.cell(row,col)
        print(cell_vlue.value)


# 先把每一列读取出来再将值赋给每一行
# for col in range(0,cols):
#     col_value=sheet2.col_values(col)
#     print(col_value)

# 逐行写入
for row in range (0,rows):
    for col in range (0,cols):
        print("当前行",row,"当前列",col)
        sheet2_write.write(row,col,sheet2.cell(row,col))
print("写入成功")


tc_wb.save('F:/Excel/测试工作表.xlsx')




"""
 行列携带数据进行交换
 idea，首先根据没一列的对应值写入每一行
 首先读取每一列
 将每一列写入每一行
 在表二sheet2中生成
"""





